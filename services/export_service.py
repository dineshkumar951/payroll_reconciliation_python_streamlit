from io import BytesIO
import pandas as pd

from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment
)

from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# SHARED STYLE HELPERS
# ---------------------------------------------------------------------------

_THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

_HEADER_FILL = PatternFill(
    start_color="D9EAF7",
    end_color="D9EAF7",
    fill_type="solid"
)

_ACCRUED_TITLE_FILL = PatternFill(
    start_color="C55A11",
    end_color="C55A11",
    fill_type="solid"
)


def _style_header_row(ws, row_num, num_cols):
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(horizontal="center")


def _style_data_rows(ws, start_row, end_row):
    for row in ws.iter_rows(min_row=start_row, max_row=end_row):
        for cell in row:
            cell.border = _THIN_BORDER
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'


def _auto_column_width(ws):
    for column_cells in ws.columns:
        length = max(
            len(str(cell.value)) if cell.value else 0
            for cell in column_cells
        )
        ws.column_dimensions[
            get_column_letter(column_cells[0].column)
        ].width = length + 5


# ---------------------------------------------------------------------------
# MAIN EXPORT
# ---------------------------------------------------------------------------

def generate_excel_file(
    grouped_df,
    checks_df,
    accrued_df=None
):

    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:

        # -------------------------------------------------------------------
        # Write data frames to sheets (startrow=2 leaves row 1 for title)
        # -------------------------------------------------------------------

        grouped_df.to_excel(
            writer,
            sheet_name="Payroll Summary",
            startrow=2,
            index=False
        )

        checks_df.to_excel(
            writer,
            sheet_name="Review Checks",
            startrow=2,
            index=False
        )

        workbook = writer.book

        # ===================================================================
        # PAYROLL SUMMARY SHEET
        # ===================================================================

        ws = workbook["Payroll Summary"]
        num_cols = len(grouped_df.columns)

        # Title row
        ws["A1"] = "Payroll Reconciliation Summary Report"
        ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
        ws["A1"].fill = PatternFill(
            start_color="1F4E78", end_color="1F4E78", fill_type="solid"
        )
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.merge_cells(
            start_row=1, start_column=1,
            end_row=1, end_column=num_cols
        )

        # Header + data rows
        _style_header_row(ws, 3, num_cols)
        normal_data_end = ws.max_row

        # -------------------------------------------------------------------
        # Accrued section — appended after normal data
        # -------------------------------------------------------------------

        if accrued_df is not None and len(accrued_df) > 0:

            accrued_num_cols = len(accrued_df.columns)

            # Blank separator row
            separator_row = ws.max_row + 2

            # "Accrued" title row
            title_row = separator_row
            ws.cell(row=title_row, column=1).value = "Accrued"
            ws.cell(row=title_row, column=1).font = Font(
                bold=True, size=14, color="FFFFFF"
            )
            ws.cell(row=title_row, column=1).fill = _ACCRUED_TITLE_FILL
            ws.cell(row=title_row, column=1).alignment = Alignment(
                horizontal="center"
            )
            ws.merge_cells(
                start_row=title_row, start_column=1,
                end_row=title_row, end_column=max(num_cols, accrued_num_cols)
            )

            # Accrued column header row
            accrued_header_row = title_row + 1
            for col_idx, col_name in enumerate(accrued_df.columns, start=1):
                cell = ws.cell(row=accrued_header_row, column=col_idx)
                cell.value = col_name
            _style_header_row(ws, accrued_header_row, accrued_num_cols)

            # Accrued data rows
            accrued_data_start = accrued_header_row + 1
            for r_idx, (_, data_row) in enumerate(
                accrued_df.iterrows(), start=accrued_data_start
            ):
                for c_idx, value in enumerate(data_row, start=1):
                    ws.cell(row=r_idx, column=c_idx).value = value

            _style_data_rows(
                ws,
                accrued_data_start,
                accrued_data_start + len(accrued_df) - 1
            )

        _style_data_rows(ws, 4, normal_data_end)
        _auto_column_width(ws)
        ws.freeze_panes = "A4"

        # ===================================================================
        # REVIEW CHECKS SHEET
        # ===================================================================

        ws2 = workbook["Review Checks"]
        num_cols2 = len(checks_df.columns)

        ws2["A1"] = "Payroll Review Checks"
        ws2["A1"].font = Font(bold=True, size=16, color="FFFFFF")
        ws2["A1"].fill = PatternFill(
            start_color="38761D", end_color="38761D", fill_type="solid"
        )
        ws2.merge_cells(
            start_row=1, start_column=1,
            end_row=1, end_column=num_cols2
        )

        _style_header_row(ws2, 3, num_cols2)
        _style_data_rows(ws2, 4, ws2.max_row)
        _auto_column_width(ws2)
        ws2.freeze_panes = "A4"

    output.seek(0)

    return output
